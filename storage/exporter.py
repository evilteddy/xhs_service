"""
Multi-format data exporter for Xiaohongshu crawler.
Supports exporting crawled note data to Excel, CSV, JSON, and Google Sheets.

@author jinbiao.sun
"""

import os
import json
import logging
from datetime import datetime
from typing import List, Dict, Any

import pandas as pd
import openpyxl

try:
    import gspread
    from google.oauth2.service_account import Credentials as ServiceAccountCredentials
    GSPREAD_AVAILABLE = True
except ImportError:
    GSPREAD_AVAILABLE = False

from utils.helpers import ensure_dir

logger = logging.getLogger(__name__)

# Column mapping: internal key -> display name
COLUMN_MAP = {
    'note_id': 'Note ID',
    'note_type': 'Note Type',
    'title': 'Title',
    'content': 'Content',
    'author': 'Author',
    'author_id': 'Author ID',
    'likes': 'Likes',
    'comments': 'Comments',
    'collects': 'Collects',
    'shares': 'Shares',
    'publish_time_str': 'Publish Time',
    'note_link': 'Note Link',
    'author_link': 'Author Link',
    'image_urls': 'Image URLs',
    'tags': 'Tags',
    'liked': 'Liked',
}


class DataExporter:
    """
    Exports crawled note data to various file formats including Google Sheets.
    """

    def __init__(
        self,
        output_dir: str = './data/exports',
        google_sheets_config: Dict[str, Any] | None = None,
    ):
        """
        Initialize the DataExporter.

        Args:
            output_dir: Directory to save exported files.
            google_sheets_config: Optional dict with keys:
                - credentials_file: Path to Google service account JSON key file.
                - spreadsheet_id: (optional) Existing spreadsheet ID to write into.
                - spreadsheet_name: (optional) Name for a new spreadsheet if no ID given.
        """
        self._output_dir = output_dir
        self._gsheet_cfg = google_sheets_config or {}
        ensure_dir(output_dir)

    def export(
        self,
        notes: List[Dict[str, Any]],
        keyword: str,
        formats: List[str] | None = None,
    ) -> List[str]:
        """
        Export notes to the specified formats.

        Args:
            notes: List of note dictionaries.
            keyword: The keyword used for crawling (used in filename).
            formats: List of format strings ('excel', 'csv', 'json').
                     Defaults to all three if not specified.

        Returns:
            List of file paths that were created.
        """
        if formats is None:
            formats = ['excel', 'csv', 'json']

        if not notes:
            logger.warning("No notes to export.")
            return []

        # Prepare DataFrame
        df = self._prepare_dataframe(notes)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        safe_keyword = keyword.replace(' ', '_').replace('/', '_')[:30]
        base_name = f"xhs_{safe_keyword}_{timestamp}"

        created_files = []
        for fmt in formats:
            fmt_lower = fmt.lower().strip()
            try:
                if fmt_lower == 'excel':
                    path = self._export_excel(df, base_name)
                elif fmt_lower == 'csv':
                    path = self._export_csv(df, base_name)
                elif fmt_lower == 'json':
                    path = self._export_json(notes, base_name)
                elif fmt_lower == 'google_sheets':
                    path = self._export_google_sheets(df, keyword)
                else:
                    logger.warning(f"Unknown export format: {fmt}")
                    continue
                created_files.append(path)
                logger.info(f"Exported {fmt_lower}: {path}")
            except Exception as e:
                logger.error(f"Failed to export {fmt_lower}: {e}")

        return created_files

    def _prepare_dataframe(self, notes: List[Dict[str, Any]]) -> pd.DataFrame:
        """
        Convert notes into a pandas DataFrame with proper formatting.

        Args:
            notes: List of note dictionaries.

        Returns:
            Cleaned and sorted DataFrame.
        """
        rows = []
        for note in notes:
            row = {}
            for key, display in COLUMN_MAP.items():
                value = note.get(key, '')
                # Convert lists to comma-separated strings for tabular formats
                if isinstance(value, list):
                    value = ', '.join(str(v) for v in value)
                row[display] = value
            rows.append(row)

        df = pd.DataFrame(rows)

        # Ensure numeric columns are integers
        for col in ['Likes', 'Comments', 'Collects']:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)

        # Remove duplicate rows
        df = df.drop_duplicates(subset=['Note ID'], keep='first')

        # Sort by comments descending
        if 'Comments' in df.columns:
            df = df.sort_values(by='Comments', ascending=False)

        df = df.reset_index(drop=True)
        return df

    def _export_excel(self, df: pd.DataFrame, base_name: str) -> str:
        """
        Export DataFrame to an Excel file with auto-adjusted column widths.

        Args:
            df: The DataFrame to export.
            base_name: Base filename (without extension).

        Returns:
            Path to the created Excel file.
        """
        filepath = os.path.join(self._output_dir, f"{base_name}.xlsx")
        df.to_excel(filepath, index=False, engine='openpyxl')

        # Auto-adjust column widths
        self._auto_resize_excel(filepath)
        return filepath

    def _export_csv(self, df: pd.DataFrame, base_name: str) -> str:
        """
        Export DataFrame to a CSV file with UTF-8 BOM encoding.

        Args:
            df: The DataFrame to export.
            base_name: Base filename (without extension).

        Returns:
            Path to the created CSV file.
        """
        filepath = os.path.join(self._output_dir, f"{base_name}.csv")
        df.to_csv(filepath, index=False, encoding='utf-8-sig')
        return filepath

    def _export_json(
        self, notes: List[Dict[str, Any]], base_name: str
    ) -> str:
        """
        Export the raw note data to a JSON file.

        Args:
            notes: List of note dictionaries (original structure preserved).
            base_name: Base filename (without extension).

        Returns:
            Path to the created JSON file.
        """
        filepath = os.path.join(self._output_dir, f"{base_name}.json")

        # Serialize datetime objects
        serializable = []
        for note in notes:
            item = {}
            for k, v in note.items():
                if isinstance(v, datetime):
                    item[k] = v.isoformat()
                else:
                    item[k] = v
            serializable.append(item)

        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(serializable, f, ensure_ascii=False, indent=2)
        return filepath

    # ------------------------------------------------------------------
    # Google Sheets export
    # ------------------------------------------------------------------

    def _get_gspread_client(self) -> 'gspread.Client':
        """
        Create an authorized gspread client using a service account JSON key.

        Returns:
            Authorized gspread.Client instance.

        Raises:
            RuntimeError: If gspread is not installed or credentials not configured.
        """
        if not GSPREAD_AVAILABLE:
            raise RuntimeError(
                "Google Sheets export requires 'gspread' and 'google-auth'. "
                "Install them with: pip install gspread google-auth"
            )

        creds_file = self._gsheet_cfg.get('credentials_file', '')
        if not creds_file or not os.path.exists(creds_file):
            raise RuntimeError(
                f"Google service account credentials file not found: '{creds_file}'. "
                "Please set 'google_sheets.credentials_file' in config.yaml to the "
                "path of your service account JSON key file."
            )

        scopes = [
            'https://www.googleapis.com/auth/spreadsheets',
            'https://www.googleapis.com/auth/drive',
        ]
        credentials = ServiceAccountCredentials.from_service_account_file(
            creds_file, scopes=scopes,
        )
        return gspread.authorize(credentials)

    def _export_google_sheets(self, df: pd.DataFrame, keyword: str) -> str:
        """
        Export DataFrame to a Google Sheets spreadsheet.

        Behaviour:
          - If `spreadsheet_id` is configured → opens that spreadsheet and
            creates/overwrites a worksheet named after the keyword + timestamp.
          - If only `spreadsheet_name` is configured → creates a brand-new
            spreadsheet with that name (shared with `share_with` email if set).

        Args:
            df: The DataFrame to export.
            keyword: The search keyword (used as worksheet name).

        Returns:
            URL of the Google Sheets spreadsheet.
        """
        gc = self._get_gspread_client()

        spreadsheet_id = self._gsheet_cfg.get('spreadsheet_id', '')
        spreadsheet_name = self._gsheet_cfg.get(
            'spreadsheet_name', 'XHS Crawler Data',
        )
        share_with_raw = self._gsheet_cfg.get('share_with', '')
        # Normalise to a list: supports both a single string and a YAML list
        if isinstance(share_with_raw, list):
            share_with = [e.strip() for e in share_with_raw if e and e.strip()]
        elif isinstance(share_with_raw, str) and share_with_raw.strip():
            share_with = [e.strip() for e in share_with_raw.split(',') if e.strip()]
        else:
            share_with = []

        timestamp = datetime.now().strftime('%m%d_%H%M')
        safe_keyword = keyword.replace('/', '_')[:20]
        worksheet_title = f"{safe_keyword}_{timestamp}"

        if spreadsheet_id:
            # Open existing spreadsheet
            sh = gc.open_by_key(spreadsheet_id)
            logger.info(f"Opened existing Google Sheet: {sh.title}")
        else:
            # Create a new spreadsheet
            sh = gc.create(spreadsheet_name)
            logger.info(f"Created new Google Sheet: '{spreadsheet_name}'")

            # Share with the user's personal Google account(s) so they can see it
            for email in share_with:
                sh.share(email, perm_type='user', role='writer')
                logger.info(f"Shared spreadsheet with: {email}")

        # Create or get a worksheet
        try:
            worksheet = sh.add_worksheet(
                title=worksheet_title,
                rows=len(df) + 1,
                cols=len(df.columns),
            )
        except gspread.exceptions.APIError:
            # Worksheet may already exist; try to get it
            worksheet = sh.worksheet(worksheet_title)
            worksheet.clear()

        # Prepare data: header row + data rows
        header = list(df.columns)
        rows = df.fillna('').astype(str).values.tolist()
        all_data = [header] + rows

        # Batch update (much faster than cell-by-cell)
        worksheet.update(all_data, value_input_option='USER_ENTERED')

        # Auto-resize / freeze header row
        try:
            worksheet.freeze(rows=1)
            worksheet.format('1:1', {'textFormat': {'bold': True}})
        except Exception as e:
            logger.debug(f"Could not format header row: {e}")

        url = sh.url
        logger.info(f"Google Sheets export complete: {url}")
        return url

    @staticmethod
    def _auto_resize_excel(filepath: str) -> None:
        """
        Auto-resize column widths in an Excel file for readability.
        - Title and Content columns: auto-fit based on content length
        - URL columns: fixed width (25)
        - Other columns: moderate auto-fit

        Args:
            filepath: Path to the Excel file.
        """
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        # Columns that contain long URLs and should be fixed-width
        url_columns = {'Note Link', 'Author Link', 'Image URLs'}
        # Columns that may have long content
        wide_columns = {'Title', 'Content', 'Tags'}

        for col_cells in ws.iter_cols(min_row=1, max_row=ws.max_row):
            header = col_cells[0].value
            col_letter = col_cells[0].column_letter

            if header in url_columns:
                ws.column_dimensions[col_letter].width = 25
            elif header in wide_columns:
                max_len = 0
                for cell in col_cells:
                    try:
                        cell_len = len(str(cell.value or ''))
                        if cell_len > max_len:
                            max_len = cell_len
                    except Exception:
                        pass
                # Cap the width at 80 characters for readability
                adjusted = min((max_len + 2) * 1.2, 80)
                ws.column_dimensions[col_letter].width = max(adjusted, 15)
            else:
                max_len = 0
                for cell in col_cells:
                    try:
                        cell_len = len(str(cell.value or ''))
                        if cell_len > max_len:
                            max_len = cell_len
                    except Exception:
                        pass
                adjusted = min((max_len + 2) * 1.2, 30)
                ws.column_dimensions[col_letter].width = max(adjusted, 10)

        wb.save(filepath)
